# -*- coding: utf-8 -*-
"""
DICOM 影像批量下载工具（可视化版）

功能：
    1. 读取 Excel 表格中的「影像号」(StudyInstanceUID) 列
    2. 通过 DICOM Q/R 协议（C-Move）从医院 PACS 批量拉取影像到本地
    3. 所有连接参数（PACS IP/端口/AE、本机 AE Title/接收端口）可视化配置

协议依据：众阳云开放服务 API「2.8.1.3 获取检查图像」——DICOM 3.0 Q/R
    C-Move 入参：StudyInstanceUID (0020,000D)

运行环境：Windows（打包为 exe 后双击运行即可，无需安装 Python）
"""

import csv
import json
import os
import queue
import socket
import sys
import threading
import time
import traceback

import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext

from pynetdicom import AE, evt
from pynetdicom.presentation import StoragePresentationContexts
try:
    # 更全的存储 SOP 类（含各类压缩格式），能接收更多推送，减少前置机"子操作失败"
    from pynetdicom.presentation import AllStoragePresentationContexts as _ALL_STORAGE_CTX
except Exception:
    _ALL_STORAGE_CTX = None
from pynetdicom.sop_class import (
    StudyRootQueryRetrieveInformationModelFind,
    StudyRootQueryRetrieveInformationModelMove,
    Verification,
)
from pydicom.dataset import Dataset


# ---------------------------------------------------------------------------
# 全局状态
# ---------------------------------------------------------------------------
OUTPUT_ROOT = ""
_store_server = None            # Store SCP 服务实例
_store_started = False
_store_aet = ""                 # Store SCP 当前监听的 AE Title
_store_port = 0                 # Store SCP 当前监听的端口
_stop_event = threading.Event()
_rate_limit_kbps = 0           # 传输限速（KB/s），0 = 不限；由批量下载开始前设置
_rate_lock = threading.Lock()
_rate_last_time = 0.0          # 令牌桶：上次补充令牌的时间
_rate_tokens = 0.0             # 令牌桶：当前可用令牌（字节）
_throttle_cap_logged = False   # 限速封顶提示是否已输出（避免刷屏）
_active_assocs = []            # 当前活跃的 DICOM association，停止时强制中断
_assoc_lock = threading.Lock()
# 当前正在下载的检查信息（供 handle_store 通知 GUI 进度用）
_current_label = ""
_current_expected_images = 0   # 前置机报告的预期影像数（用于显示"X / N"）
_download_start_time = 0.0     # 本次批量下载的开始时间戳（用于"已耗时"）
_diag_log = False              # 诊断日志开关：PDU/协商/Pending 细节默认隐藏，排查时打开
# 反向探测状态：检测连通时发空 C-Move，观察前置机是否主动连入本机接收服务
_probe_active = False
_probe_conn_event = threading.Event()
_probe_conn_addr = ""

# 下载结果记录（每次下载一条，最终写入 CSV 报告并在 GUI 历史列表展示）
_download_records = []
_records_lock = threading.Lock()

# 线程间通信队列（工作线程 -> GUI 主线程）
_ui_queue = queue.Queue()


def _diag(msg):
    """诊断日志：仅在打开「诊断日志」开关时输出（PDU/协商/Pending 等细节）。"""
    if _diag_log:
        _ui_queue.put(("log", msg))


def get_app_dir():
    """返回程序所在目录（兼容 PyInstaller 打包后的 exe）。"""
    if getattr(sys, "frozen", False):
        return os.path.dirname(sys.executable)
    return os.path.dirname(os.path.abspath(__file__))


CONFIG_PATH = os.path.join(get_app_dir(), "config.json")


# ---------------------------------------------------------------------------
# 1) Store SCP：接收 PACS 通过 C-Move 推送的 DICOM 文件
# ---------------------------------------------------------------------------
def _safe_name(s):
    if s is None:
        return "unnamed"
    s = str(s)
    safe = "".join(c if (c.isalnum() or c in "._-") else "_" for c in s)
    # 极端兜底：纯特殊字符（如 NULL 字节）会被全部替换为 _，但仍然是非空
    return safe if safe.strip("_") else "unnamed"


def _on_conn_open(event):
    """Store SCP 收到入站连接时记录来源，便于判断 PACS 是否真的在往本机推影像。"""
    global _probe_conn_addr
    try:
        addr = "%s:%s" % (event.address[0], event.address[1])
    except Exception:
        addr = "未知"
    _ui_queue.put(("log", "[StoreSCP] 收到入站连接：%s" % addr))
    if _probe_active:
        _probe_conn_addr = addr
        _probe_conn_event.set()


def _on_conn_close(event):
    """Store SCP 连接关闭时记录来源，配合「入站连接」判断对方连上后做了什么。"""
    try:
        addr = "%s:%s" % (event.address[0], event.address[1])
    except Exception:
        addr = "未知"
    _ui_queue.put(("log", "[StoreSCP] 连接关闭：%s" % addr))


def _on_echo(event):
    """Store SCP 支持 C-ECHO：部分前置机推影像前会先对目标 AE 做 C-ECHO 校验，
    不支持会导致对方中止推送（表现为 C-Move 成功但 0 文件落盘）。"""
    try:
        addr = "%s:%s" % (event.assoc.remote_address[0], event.assoc.remote_address[1])
    except Exception:
        addr = "未知"
    _ui_queue.put(("log", "[StoreSCP] 收到 C-ECHO 校验：%s（已应答成功）" % addr))
    return 0x0000


def _ae_str(v):
    """AE Title 可能是 bytes，统一转成可读字符串。"""
    try:
        if isinstance(v, bytes):
            return v.decode("ascii", "replace").strip()
        return str(v)
    except Exception:
        return "?"


def _cx_abstract_name(cx):
    try:
        return cx.abstract_syntax.name
    except Exception:
        try:
            return str(cx.abstract_syntax)
        except Exception:
            return "?"


def _cx_transfer_name(cx):
    try:
        ts = cx.transfer_syntax
        if isinstance(ts, (list, tuple)):
            ts = ts[0] if ts else ""
        return _ae_str(ts)
    except Exception:
        return "?"


def _pdu_summary(pdu):
    """PDU 一行摘要。P-DATA-TF（影像数据块）只输出字节长度，避免刷屏。"""
    name = type(pdu).__name__
    if name == "P_DATA_TF":
        try:
            n = pdu.pdu_length
        except Exception:
            n = -1
        return "P-DATA-TF（数据块 %d 字节）" % n
    return name


def _make_pdu_handlers(side):
    """带侧别前缀的 PDU 捕获处理器（诊断级，默认不输出），避免 SCU/SCP 两侧日志混淆。"""

    def _recv(event):
        try:
            _diag("[%s][PDU←] %s" % (side, _pdu_summary(event.pdu)))
        except Exception:
            pass

    def _sent(event):
        try:
            _diag("[%s][PDU→] %s" % (side, _pdu_summary(event.pdu)))
        except Exception:
            pass

    return _recv, _sent


def _log_assoc_negotiation(assoc, tag):
    """输出关联协商结果（双方 AE、同意/拒绝的呈现上下文）。诊断级，默认不输出。"""
    try:
        remote = "%s:%s" % (assoc.remote_address[0], assoc.remote_address[1])
    except Exception:
        remote = "未知"
    try:
        calling = _ae_str(assoc.requestor.primitive.calling_ae_title)
        called = _ae_str(assoc.requestor.primitive.called_ae_title)
    except Exception:
        calling, called = "?", "?"
    _diag("[%s] 关联协商完成：对端=%s 调用方AE=%s 被叫方AE=%s" % (tag, remote, calling, called))
    try:
        for cx in assoc.accepted_contexts:
            _diag("[%s]   [同意] %s / %s" % (tag, _cx_abstract_name(cx), _cx_transfer_name(cx)))
        for cx in assoc.rejected_contexts:
            _diag("[%s]   [拒绝] %s" % (tag, _cx_abstract_name(cx)))
    except Exception:
        pass


def _on_accepted(event):
    """Store SCP 接受关联时输出协商结果：排查"对方连上却不发 C-STORE"。"""
    _log_assoc_negotiation(event.assoc, "StoreSCP协商")


def _on_aborted(event):
    """关联被 ABORT 时记录来源：区分"对方主动中止"与"正常释放"。"""
    try:
        addr = "%s:%s" % (event.assoc.remote_address[0], event.assoc.remote_address[1])
    except Exception:
        addr = "未知"
    _ui_queue.put(("log", "[StoreSCP] 关联被中止（A-ABORT）：%s" % addr))


def _count_dcm(subdir):
    try:
        return len([f for f in os.listdir(subdir) if f.endswith(".dcm")])
    except Exception:
        return 0


def _local_ips():
    """本机非回环 IPv4 地址列表，用于核对 PACS 侧注册的 IP。"""
    ips = []
    try:
        for info in socket.getaddrinfo(socket.gethostname(), None, socket.AF_INET):
            ip = info[4][0]
            if ip not in ips:
                ips.append(ip)
    except Exception:
        pass
    return ips


def handle_store(event):
    """处理 C-STORE：把收到的 DICOM 文件保存到对应 Study 子目录（并发安全）。"""
    try:
        ds = event.dataset
        ds.file_meta = event.file_meta
        sop_uid = getattr(ds, "SOPInstanceUID", None) or ("unknown-%d" % int(time.time()))
        study_uid = getattr(ds, "StudyInstanceUID", None)
        # 关键防御：OUTPUT_ROOT 尚未设置或已被清空时（如程序关闭中/已停止）拒绝写入
        if not OUTPUT_ROOT:
            _ui_queue.put(("log", "[StoreSCP] 收到 DICOM 但输出目录未就绪，已拒绝：SOP=%s" % sop_uid))
            return 0xC000
        # 统一优先用图像自带的 PatientID 作目录名（缺失时回落到 StudyInstanceUID）
        patient_id = getattr(ds, "PatientID", None)
        folder_key = patient_id or study_uid
        if folder_key:
            d = os.path.join(OUTPUT_ROOT, _safe_name(folder_key))
        else:
            d = os.path.join(OUTPUT_ROOT, "_unknown_study")
            _ui_queue.put(("log", "[StoreSCP] 收到缺少定位字段的文件，落入 _unknown_study/"))
        os.makedirs(d, exist_ok=True)
        path = os.path.join(d, _safe_name(sop_uid) + ".dcm")
        # 极少数情况下（pynetdicom 内部并发回调）save_as 可能撞到 Windows 文件锁
        # 一次重试即可恢复
        try:
            ds.save_as(path, enforce_file_format=True)
        except Exception as e:
            _ui_queue.put(("log", "[StoreSCP] 首次保存失败，重试一次：%s" % e))
            time.sleep(0.05)
            ds.save_as(path, enforce_file_format=True)
        # 限速：按实际落盘文件大小节流，压低下行速率
        try:
            _throttle(os.path.getsize(path))
        except Exception:
            pass
        # 通知 GUI：当前检查已收到一个影像
        try:
            n_in_dir = _count_dcm(d)
            # 分母用累计最大值：handle_store 是异步回调，可能在 pull_one_study 下一次更新
            # _current_expected_images 之前就读到比 n_in_dir 小的旧值，导致瞬间「分子 > 分母」
            n_total = max(n_in_dir, _current_expected_images)
            _ui_queue.put(("image_progress", (_current_label, n_in_dir, n_total)))
        except Exception:
            pass
        return 0x0000  # Success
    except Exception as e:
        _ui_queue.put(("log", "[StoreSCP] 保存 DICOM 失败：%s" % e))
        return 0xC000  # Unable to process，避免单文件失败中断接收


def _throttle(size_bytes):
    """全局限速（并发安全，令牌桶）：把多个 C-STORE 的总下行速率压到目标值。

    注意：限速通过"延迟 C-STORE 应答"实现，前置机若长时间收不到应答会把该子操作
    记为失败（表现为 failed=N、最终 0xA702）。因此单文件等待封顶 2 秒：
    限速值过低时实际速率会高于设定值，但保证下载不被前置机超时打断。
    """
    global _rate_last_time, _rate_tokens, _throttle_cap_logged
    limit = _rate_limit_kbps
    if limit <= 0 or size_bytes <= 0:
        return
    rate = limit * 1024.0  # 字节/秒
    waited = 0.0
    while True:
        with _rate_lock:
            now = time.time()
            if _rate_last_time == 0:
                _rate_last_time = now
                _rate_tokens = rate  # 初始满桶（约 1 秒额度）
            else:
                elapsed = now - _rate_last_time
                _rate_tokens = min(rate, _rate_tokens + elapsed * rate)
                _rate_last_time = now
            if _rate_tokens >= size_bytes:
                _rate_tokens -= size_bytes
                return
            need = (size_bytes - _rate_tokens) / rate
        # 单文件等待封顶 2 秒：避免 C-STORE 应答过慢被前置机判超时（failed/0xA702）
        if waited >= 2.0:
            if not _throttle_cap_logged:
                _throttle_cap_logged = True
                _ui_queue.put(("log", "[限速] 限速值过低，单文件等待已封顶 2 秒，实际速率将高于设定值（不影响下载成功）"))
            return
        step = min(need, 5.0, 2.0 - waited)
        time.sleep(step)
        waited += step


def _try_bind_port(port):
    """尝试绑定本机端口（绑定后立即释放），用于检测端口是否可用。返回 (ok, err)。"""
    s = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
    try:
        s.setsockopt(socket.SOL_SOCKET, socket.SO_REUSEADDR, 1)
        s.bind(("0.0.0.0", port))
        return True, ""
    except OSError as e:
        return False, str(e)
    finally:
        try:
            s.close()
        except Exception:
            pass


def _port_bind_hint(port):
    """本机端口无法绑定时的可操作建议。"""
    return ("端口 %d 无法绑定的可能原因：\n"
            "1) 本工具的其它实例或其它程序正占用该端口，请关闭后重试；\n"
            "2) 端口落在 Windows Hyper-V/WSL 保留的排除范围内，可在 cmd 运行 "
            "netsh interface ipv4 show excludedportrange protocol=tcp 查看，"
            "改用一个不在范围内的端口（需重新向医院前置机注册）或重启电脑后重试；\n"
            "3) 被杀毒软件/防火墙拦截，请将本程序加入白名单。" % port)


def start_store_scp(local_aet, local_port):
    """启动本机 Store SCP（后台线程）。AE/端口变化时自动重启，保证与 C-Move 目标一致。"""
    global _store_server, _store_started, _store_aet, _store_port
    if _store_started and _store_server is not None:
        if _store_aet == local_aet and _store_port == local_port:
            _ui_queue.put(("log", "[StoreSCP] 接收服务已运行（AE=%s 端口=%d）" % (local_aet, local_port)))
            return True
        # 监听配置变化：关闭旧服务，重新启动
        _ui_queue.put(("log", "[StoreSCP] 本地接收参数变化，正在重启（AE=%s 端口=%d）..." % (local_aet, local_port)))
        try:
            _store_server.shutdown()
        except Exception as e:
            _ui_queue.put(("log", "[StoreSCP] 关闭旧接收服务异常（忽略）：%s" % e))
        _store_server = None
        _store_started = False
        # 端口释放后等待再 bind，避免 TIME_WAIT 导致新服务启动失败
        time.sleep(0.3)

    ae = AE()
    # 优先用更全的存储上下文（含压缩格式），减少前置机因"格式不被接收"而子操作失败
    storage_ctxs = _ALL_STORAGE_CTX or StoragePresentationContexts
    # StoragePresentationContexts 是列表，需要逐个 add_supported_context
    for ctx in storage_ctxs:
        ae.add_supported_context(ctx.abstract_syntax, ctx.transfer_syntax)
    ae.add_requested_context(Verification)
    # 作为 SCP 也要支持 C-ECHO：部分前置机推影像前先对目标 AE 做 C-ECHO 校验
    ae.add_supported_context(Verification)
    pdu_recv, pdu_sent = _make_pdu_handlers("SCP")
    handlers = [
        (evt.EVT_C_STORE, handle_store),
        (evt.EVT_C_ECHO, _on_echo),
        (evt.EVT_CONN_OPEN, _on_conn_open),
        (evt.EVT_CONN_CLOSE, _on_conn_close),
        (evt.EVT_ACCEPTED, _on_accepted),
        (evt.EVT_ABORTED, _on_aborted),
        (evt.EVT_PDU_RECV, pdu_recv),
        (evt.EVT_PDU_SENT, pdu_sent),
    ]
    try:
        ae.ae_title = local_aet
    except Exception:
        pass
    try:
        _store_server = ae.start_server(
            ("0.0.0.0", local_port), block=False, evt_handlers=handlers, ae_title=local_aet,
        )
    except Exception as e:
        _ui_queue.put(("error", "Store SCP 启动失败：%s\n%s" % (e, _port_bind_hint(local_port))))
        return False
    _store_started = True
    _store_aet = local_aet
    _store_port = local_port
    ips = "、".join(_local_ips()) or "未知"
    _ui_queue.put(("log", "[StoreSCP] 接收服务已启动：AE=%s 端口=%d（本机 IP：%s，请确认与前置机注册一致）" % (local_aet, local_port, ips)))
    return True


# ---------------------------------------------------------------------------
# 2) 网络可达性检测
# ---------------------------------------------------------------------------
def test_connectivity(host, port, aet=None, local_port=None, local_aet=None):
    """
    双向检测 PACS 前置机与本机的连通性。
    返回 (ok, messages)：
        正向（本机 -> 前置机）：
            1. TCP 端口连通性检查
            2. DICOM C-Echo 检查（验证前置机服务与 AE Title，需要提供 aet）
        反向（前置机 -> 本机）：
            3. 本机接收服务自检：经本机外部 IP 连接收服务并 C-Echo（需要 local_port/local_aet）
            4. 前置机反向探测：发一个不存在 UID 的 C-Move，观察前置机是否主动连入本机
               接收服务（仅作参考提示，不影响 ok 判定，因为部分前置机对 0 匹配不连目标）
        本机接收端口自检：启动 Store SCP 本身即端口绑定测试
    """
    global _probe_active, _probe_conn_addr
    msgs = []
    tcp_ok = False
    echo_ok = False
    local_ok = True
    self_ok = True

    # 0) 确保本机接收服务在运行（反向检测依赖它；启动本身即端口绑定自检）
    store_ready = False
    if local_port:
        if local_aet:
            store_ready = start_store_scp(local_aet, local_port)
            if store_ready:
                msgs.append("本机接收服务就绪：AE=%s 端口=%d" % (local_aet, local_port))
            else:
                local_ok = False
                msgs.append("本机接收服务启动失败：端口 %d 无法绑定" % local_port)
                msgs.append(_port_bind_hint(local_port))
        elif not (_store_started and _store_port == local_port):
            ok, err = _try_bind_port(local_port)
            if ok:
                msgs.append("本机接收端口自检：端口 %d 可正常绑定" % local_port)
            else:
                local_ok = False
                msgs.append("本机接收端口自检失败：端口 %d 无法绑定（%s）" % (local_port, err))
                msgs.append(_port_bind_hint(local_port))

    # 1) TCP 连通性（正向）
    try:
        sock = socket.create_connection((host, port), timeout=5)
        sock.close()
        tcp_ok = True
        msgs.append("TCP 连接成功：%s:%d 可达" % (host, port))
    except Exception as e:
        msgs.append("TCP 连接失败：%s:%d 不可达（%s）" % (host, port, e))
        msgs.append("请检查：IP/端口是否正确、网络是否连通、是否需 VPN/白名单")
        return False, msgs

    # 2) DICOM C-Echo（正向：验证前置机服务与 AE Title）
    if not aet:
        msgs.append("未填写 PACS AE Title，跳过正向 DICOM C-Echo 验证")
    else:
        try:
            ae = AE()
            ae.add_requested_context(Verification)
            ae.acse_timeout = 10
            ae.network_timeout = 10
            assoc = ae.associate(host, port, ae_title=aet)
            if assoc.is_established:
                try:
                    status = assoc.send_c_echo()
                    if status and getattr(status, "Status", None) == 0x0000:
                        echo_ok = True
                        msgs.append("正向 DICOM C-Echo 成功：AE Title=%s 有效" % aet)
                    else:
                        msgs.append("正向 DICOM C-Echo 未返回成功状态")
                finally:
                    try:
                        assoc.release()
                    except Exception:
                        pass
            else:
                msgs.append("正向 DICOM 关联建立失败：请检查 AE Title=%s 是否正确" % aet)
                # 未建立时也要尝试 release 释放 socket 资源
                try:
                    assoc.release()
                except Exception:
                    pass
        except Exception as e:
            msgs.append("正向 DICOM C-Echo 异常：%s" % e)

    # 3) 反向自检：经本机外部 IP 连自己的接收服务并 C-Echo（验证监听/AE/应答全链路）
    # 逐个外部 IP 尝试（VPN/虚拟网卡可能不可路由，任一成功即视为通过）
    if store_ready or (_store_started and local_port and _store_port == local_port):
        self_ok = False
        self_err = ""
        for self_ip in (_local_ips() or ["127.0.0.1"]):
            try:
                ae = AE()
                ae.add_requested_context(Verification)
                ae.acse_timeout = 5
                ae.network_timeout = 5
                assoc = ae.associate(self_ip, local_port, ae_title=local_aet or "SELFECHO")
                if assoc.is_established:
                    try:
                        st = assoc.send_c_echo()
                        if st and getattr(st, "Status", None) == 0x0000:
                            msgs.append("反向自检成功：经 %s:%d 可达本机接收服务且 C-Echo 应答正常" % (self_ip, local_port))
                            self_ok = True
                        else:
                            self_err = "C-Echo 应答异常"
                    finally:
                        try:
                            assoc.release()
                        except Exception:
                            pass
                else:
                    self_err = "无法建立关联"
                    try:
                        assoc.release()
                    except Exception:
                        pass
            except Exception as e:
                self_err = str(e)
            if self_ok:
                break
        if not self_ok:
            msgs.append("反向自检失败（%s）：请检查 Windows 防火墙是否放行端口 %d 入站" % (self_err, local_port))

    # 4) 前置机反向探测：发不存在 UID 的 C-Move，观察前置机是否主动连入本机（仅提示）
    if (store_ready or (_store_started and local_port and _store_port == local_port)) and aet and local_aet:
        _probe_conn_event.clear()
        _probe_conn_addr = ""
        _probe_active = True
        try:
            ae = AE()
            ae.add_requested_context(StudyRootQueryRetrieveInformationModelMove)
            ae.acse_timeout = 10
            ae.dimse_timeout = 5
            ae.network_timeout = 10
            assoc = ae.associate(host, port, ae_title=aet)
            if assoc.is_established:
                try:
                    ds = Dataset()
                    ds.QueryRetrieveLevel = "STUDY"
                    ds.StudyInstanceUID = "1.2.3.999.999999.999999999"  # 不存在的 UID
                    for _st, _id in assoc.send_c_move(ds, local_aet, StudyRootQueryRetrieveInformationModelMove):
                        break  # 只需触发前置机动作，取首个响应即退出
                except Exception:
                    pass
                finally:
                    try:
                        assoc.release()
                    except Exception:
                        pass
                if _probe_conn_event.wait(8):
                    msgs.append("反向探测成功：前置机已主动连入本机接收服务（%s），反向通道正常" % _probe_conn_addr)
                else:
                    ips = "、".join(_local_ips()) or "未知"
                    msgs.append("反向探测：8 秒内前置机未连入本机接收服务。"
                                "若下载仍 0 文件，请核对前置机上「%s」注册的 IP/端口是否为本机（%s / %d）"
                                % (local_aet, ips, local_port))
            else:
                msgs.append("反向探测跳过：无法与 PACS 建立关联")
                try:
                    assoc.release()
                except Exception:
                    pass
        except Exception as e:
            msgs.append("反向探测异常：%s" % e)
        finally:
            _probe_active = False

    return (tcp_ok and local_ok and self_ok and (echo_ok if aet else True)), msgs


# ---------------------------------------------------------------------------
# 3) 读取 Excel 中的「影像号」(StudyInstanceUID) 列
# ---------------------------------------------------------------------------
def read_study_uids(excel_path, column, sheet_name=None, fallback_to_first_col=True):
    """
    读取 Excel 指定列，返回去重后的非空 StudyInstanceUID 列表。
    column 可以是：
        - 列名（如 "影像号"，按表头匹配）
        - 字母（如 "A"）
        - 数字（如 "1"，1-based）
    fallback_to_first_col=False 时，按列名匹配失败会直接抛 ValueError（不再静默读第一列）。
    """
    import openpyxl
    try:
        wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    except PermissionError:
        raise PermissionError(
            "无法读取 Excel（文件被占用）：%s\n请先关闭 Excel 后再试。" % excel_path
        )
    except Exception as e:
        raise Exception("加载 Excel 失败：%s（路径：%s）" % (e, excel_path))
    try:
        ws = wb[sheet_name] if sheet_name else wb.active
    except KeyError:
        names = "、".join(wb.sheetnames)
        wb.close()
        raise ValueError("Sheet「%s」不存在，当前工作簿的 Sheet 为：%s" % (sheet_name, names))

    rows = ws.iter_rows(values_only=True)
    header = next(rows, None)
    all_rows = list(rows)

    def col_letter_to_idx(letter):
        idx = 0
        for ch in letter.upper():
            idx = idx * 26 + (ord(ch) - ord("A") + 1)
        return idx - 1

    col_idx = None

    c = str(column).strip()
    if c.isdigit():
        col_idx = int(c) - 1
    elif c and c.isascii() and c.isalpha() and len(c) <= 3:
        # 仅纯英文字母（A-Z/a-z）才按“列字母”解析，避免中文列名被误判
        col_idx = col_letter_to_idx(c)
    elif header:
        # 按表头匹配列名
        for i, h in enumerate(header):
            if str(h).strip() == c:
                col_idx = i
                break
        if col_idx is None:
            if not fallback_to_first_col:
                headers = [str(h).strip() for h in header if h is not None and str(h).strip()]
                raise ValueError(
                    "在表头中找不到列「%s」。当前表头为：%s。请修改「关键列名」，或改用列字母（如 A）/序号（如 1）"
                    % (c, "、".join(headers[:15]) if headers else "(空)")
                )
            # 表头没匹配到，退回第一列，并把表头行也当作数据
            col_idx = 0
            all_rows = [header] + all_rows
    else:
        col_idx = 0

    if col_idx is None or col_idx < 0:
        col_idx = 0

    uids = []
    seen = set()
    for r in all_rows:
        val = r[col_idx] if (r and len(r) > col_idx) else None
        if val is None:
            continue
        if isinstance(val, float) and val.is_integer():
            # 长数字（影像号/病人ID）被 Excel 存成浮点数时，避免 str() 变成科学计数法
            s = str(int(val))
        else:
            s = str(val).strip()
        if not s:
            continue
        if s not in seen:
            seen.add(s)
            uids.append(s)

    wb.close()
    return uids


# ---------------------------------------------------------------------------
# 3) C-Move 拉取单个 Study
# ---------------------------------------------------------------------------
_PENDING = {0xFF00, 0xFF01}


def _status_text(code):
    """把 DICOM C-Find/C-Move 状态码转成可读文本。"""
    m = {
        0x0000: "成功(Success)",
        0xFF00: "进行中(Pending)",
        0xFF01: "进行中(Pending, 有警告)",
        0xA700: "拒绝:资源不足(Out of Resources)",
        0xA701: "拒绝:无法计算匹配数(Unable to calculate matches)",
        0xA702: "拒绝:资源不足-无法计算匹配数",
        0xA801: "拒绝:目标节点未知(Move Destination unknown)",
        0xA900: "失败:标识符与SOP类不匹配(Identifier does not match SOP Class)",
        0xC000: "失败:无法处理(Unable to process)",
        0xC001: "失败:无法处理-部分键(Unable to process, some keys)",
        0xFE00: "取消(Cancel)",
        0xB000: "警告:子操作完成但有失败(Warning, some sub-ops failed)",
    }
    return m.get(code, "未知(0x%04X)" % code)


def find_studies_by_patient(assoc, patient_id):
    """C-Find 查询某个病人的所有 Study，返回 StudyInstanceUID 列表。"""
    ds = Dataset()
    ds.QueryRetrieveLevel = "STUDY"
    ds.PatientID = patient_id
    ds.StudyInstanceUID = ""
    ds.StudyDate = ""
    ds.Modality = ""
    ds.StudyDescription = ""
    ds.AccessionNumber = ""
    ds.StudyID = ""

    _ui_queue.put(("log", "      [C-Find] 查询 patientId=%s (QueryRetrieveLevel=STUDY)" % patient_id))
    uids = []
    seen_uids = set()  # 去重 set，避免 O(n²) 遍历
    last_status = None
    try:
        # send_c_find(dataset, query_model) - query_model 必须位置参数
        for status, identifier in assoc.send_c_find(ds, StudyRootQueryRetrieveInformationModelFind):
            if status:
                last_status = status.Status
                _ui_queue.put(("log", "      [C-Find] 响应状态 0x%04X %s" % (status.Status, _status_text(status.Status))))
            if identifier:
                suid = getattr(identifier, "StudyInstanceUID", None)
                acc = getattr(identifier, "AccessionNumber", None)
                sid = getattr(identifier, "StudyID", None)
                sdate = getattr(identifier, "StudyDate", None)
                mod = getattr(identifier, "Modality", None)
                desc = getattr(identifier, "StudyDescription", None)
                _ui_queue.put(("log", "          -> StudyInstanceUID=%s 日期=%s 模态=%s 描述=%s 申请号=%s StudyID=%s" % (
                    suid, sdate, mod, desc, acc, sid)))
                if suid:
                    suid_str = str(suid)
                    if suid_str not in seen_uids:
                        seen_uids.add(suid_str)
                        uids.append(suid_str)
    except Exception as e:
        _ui_queue.put(("log", "      [C-Find 异常] %s" % e))
    _ui_queue.put(("log", "      [C-Find] 结束，最终状态 0x%04X %s，命中 %d 个 Study" % (
        last_status or 0, _status_text(last_status) if last_status is not None else "无响应", len(uids))))
    return uids


def pull_one_study(assoc, study_uid, local_aet, out_root, folder_key=None):
    """拉取一个 Study，影像落到 out_root/<folder_key>/ 目录（folder_key 缺省用 study_uid）。

    C-Move 进行中（Pending 状态）会实时用「已完成 + 失败 + 警告」估算前置机已确定的子操作数，
    累计取最大值（避免某些前置机实现不规范导致分母倒退），并立即更新全局 _current_expected_images，
    让 handle_store 后续发的 image_progress 也能用上新分母；前置机不报子操作计数时退化为 C-Move
    完成后再用 n_files + n_failed + n_warned 兜底作为分母。

    注意：NumberOfRemainingSuboperations 在该前置机上不可信（剩余=0 但已完成还在涨），
    所以不参与估算，仅在日志中展示。
    """
    global _current_expected_images
    folder = folder_key or study_uid
    subdir = os.path.join(out_root, _safe_name(folder))
    os.makedirs(subdir, exist_ok=True)

    ds = Dataset()
    ds.QueryRetrieveLevel = "STUDY"
    ds.StudyInstanceUID = study_uid

    _ui_queue.put(("log", "      [C-Move] 请求拉取 StudyInstanceUID=%s -> 目标AE=%s (QueryRetrieveLevel=STUDY)" % (study_uid, local_aet)))

    final_code = 0x0000
    has_error = False
    n_failed = 0    # 子操作失败数（前置机推送失败的影像数）
    n_warned = 0    # 子操作警告数
    expected_total_max = 0  # 累计最大的"前置机应推总数"估算（取最大值避免分母倒退）
    try:
        # send_c_move(dataset, move_aet, query_model) - query_model 必须位置参数
        for status, _ in assoc.send_c_move(ds, local_aet, StudyRootQueryRetrieveInformationModelMove):
            if status:
                code = status.Status
                # 子操作计数（C-Move 的 Pending 阶段会带这些值，可用来评估进度）
                remaining = getattr(status, "NumberOfRemainingSuboperations", None)
                completed = getattr(status, "NumberOfCompletedSuboperations", None)
                failed = getattr(status, "NumberOfFailedSuboperations", None)
                warned = getattr(status, "NumberOfWarningSuboperations", None)
                if failed is not None:
                    n_failed = int(failed)
                if warned is not None:
                    n_warned = int(warned)
                # 估算前置机已确定要推的总数：只用 已完成 + 失败 + 警告
                # 注意：NumberOfRemainingSuboperations 在该前置机上不可信（剩余=0 但已完成还在涨，
                # 说明它只在前几个 Pending 报一次或干脆不更新；强行参与估算会导致分母小于分子）
                if completed is not None or failed is not None or warned is not None:
                    c_val = int(completed) if completed is not None else 0
                    f_val = int(failed) if failed is not None else 0
                    w_val = int(warned) if warned is not None else 0
                    est_from_subops = c_val + f_val + w_val
                    # 取「子操作估算值」与「当前已落盘数」的较大者，防止前置机的"已完成"上报有延迟
                    # 时显示「分子 > 分母」；同时永不倒退
                    n_files_now = _count_dcm(subdir)
                    est = max(est_from_subops, n_files_now, expected_total_max)
                    if est > expected_total_max:
                        expected_total_max = est
                        # 实时更新全局分母：handle_store 后续发的 image_progress 会立即用上新分母
                        _current_expected_images = est
                        # 主动推一次进度，让 UI 立即从「X / ?」变成「X / N」
                        _ui_queue.put(("image_progress", (_current_label, n_files_now, est)))
                extra = ""
                if remaining is not None or completed is not None:
                    extra = " (剩余=%s 已完成=%s 失败=%s 警告=%s)" % (remaining, completed, failed, warned)
                msg = "      [C-Move] 响应状态 0x%04X %s%s" % (code, _status_text(code), extra)
                if code in _PENDING:
                    _diag(msg)  # Pending 过程日志默认隐藏，避免刷屏
                else:
                    _ui_queue.put(("log", msg))
                if code not in _PENDING:
                    final_code = code
                    # 0xB000=警告:子操作完成但有失败，属于“部分成功”，不当作硬错误；
                    # 由 _download_one 按本地实际落盘文件数记为成功，接受个别失败
                    if code not in (0x0000, 0xB000):
                        has_error = True
    except Exception as e:
        has_error = True
        _ui_queue.put(("log", "      [C-Move 异常] %s" % e))

    n_files = _count_dcm(subdir)
    # 部分医院前置机是"先返回 C-Move 成功、再异步推送影像"，
    # 状态成功但暂时 0 文件时，最多等 30 秒观察文件是否陆续落盘
    if not has_error and n_files == 0:
        _ui_queue.put(("log", "      [C-Move] 状态成功但暂无文件落盘，等待前置机异步推送（最多 30 秒）..."))
        for _ in range(60):
            time.sleep(0.5)
            if _stop_event.is_set():
                break
            n_files = _count_dcm(subdir)
            if n_files > 0:
                break
    _ui_queue.put(("log", "      [C-Move] 结束，最终状态 0x%04X %s，本地落盘 %d 个文件" % (final_code, _status_text(final_code), n_files)))
    # n_total 兜底：取 (已落盘 + 已失败 + 已警告)、(累计最大预期值)、(已落盘数) 三者最大
    # 防止前置机"剩余=0 但还在异步推"导致最终分母小于实际收到的影像数
    n_total = max(n_files + n_failed + n_warned, expected_total_max, n_files)
    return has_error, final_code, n_files, subdir, n_failed, n_warned, n_total


# ---------------------------------------------------------------------------
# 4) 批量下载（后台线程执行）
# ---------------------------------------------------------------------------
class DownloadConfig:
    def __init__(self):
        self.pacs_host = ""
        self.pacs_port = 104
        self.pacs_aet = ""
        self.local_aet = "MYAET"
        self.local_port = 11112
        self.excel_path = ""
        self.sheet_name = ""
        self.column = "影像号"
        self.key_type = "patient_id"  # patient_id(病人ID) / study_uid(StudyInstanceUID)，默认按病人ID
        self.out_dir = ""
        self.rate_limit_kbps = 0      # 限速（KB/s），0 = 不限
        self.pause_every = 0          # 每下载 N 个检查后暂停（仅串行模式），0 = 不启用
        self.pause_seconds = 30       # 暂停秒数
        self.cmove_timeout = 300      # C-MOVE 超时（秒）
        self.cfind_timeout = 60       # C-FIND 超时（秒）


def _register_assoc(assoc):
    with _assoc_lock:
        _active_assocs.append(assoc)


def _unregister_assoc(assoc):
    with _assoc_lock:
        try:
            _active_assocs.remove(assoc)
        except ValueError:
            pass


def _abort_active_assocs():
    """停止时强制中断所有活跃的 association，让卡在网络等待上的线程尽快退出。"""
    with _assoc_lock:
        for a in list(_active_assocs):
            try:
                a.abort()
            except Exception:
                pass


def _make_assoc(cfg, sop_class, label, timeout_attr, timeout_default):
    """建立 DICOM association（C-Find / C-Move 共用）。连接异常返回 None。"""
    _ui_queue.put(("log", "正在连接 PACS %s:%d（%s，AE=%s）..." % (cfg.pacs_host, cfg.pacs_port, label, cfg.pacs_aet)))
    ae = AE()
    ae.add_requested_context(sop_class)
    ae.acse_timeout = 15
    ae.dimse_timeout = int(getattr(cfg, timeout_attr, timeout_default) or timeout_default)
    ae.network_timeout = 30
    # 诊断：SCU 侧关联协商结果 + 底层 PDU 流捕获
    pdu_recv, pdu_sent = _make_pdu_handlers("SCU")
    handlers = [
        (evt.EVT_ACCEPTED, _on_accepted_scu),
        (evt.EVT_ABORTED, _on_aborted_scu),
        (evt.EVT_PDU_RECV, pdu_recv),
        (evt.EVT_PDU_SENT, pdu_sent),
    ]
    try:
        assoc = ae.associate(cfg.pacs_host, cfg.pacs_port, ae_title=cfg.pacs_aet, evt_handlers=handlers)
    except Exception as e:
        _ui_queue.put(("log", "[%s] 连接 PACS 异常：%s" % (label, e)))
        return None
    if assoc.is_established:
        _register_assoc(assoc)
    else:
        # 关联未建立（被 PACS 拒、网络异常等），确保释放对象内部 socket 资源
        try:
            assoc.release()
        except Exception:
            pass
    return assoc


def _on_accepted_scu(event):
    _log_assoc_negotiation(event.assoc, "SCU协商")


def _on_aborted_scu(event):
    try:
        addr = "%s:%s" % (event.assoc.remote_address[0], event.assoc.remote_address[1])
    except Exception:
        addr = "未知"
    _ui_queue.put(("log", "[SCU] 关联被中止（A-ABORT）：%s" % addr))


def _record_download(idx, key, uid, label, status, n_files, message=""):
    """追加一条下载结果记录（线程安全）。"""
    rec = {
        "idx": idx,
        "key": key,
        "study_uid": uid,
        "label": label,
        "status": status,
        "n_files": n_files,
        "message": message,
        "time": time.strftime("%Y-%m-%d %H:%M:%S"),
    }
    with _records_lock:
        _download_records.append(rec)


def _clear_records():
    with _records_lock:
        _download_records.clear()


def _snapshot_records():
    with _records_lock:
        return list(_download_records)


_REPORT_HEADER = ["序号", "查询键", "StudyInstanceUID", "结果", "文件数", "说明", "时间"]


def _write_report_csv(records, out_root):
    """把下载记录写入输出目录下的 CSV 报告，返回文件路径；失败返回 None。"""
    if not out_root:
        return None
    try:
        os.makedirs(out_root, exist_ok=True)
    except Exception:
        return None
    csv_path = os.path.join(out_root, "下载报告_%s.csv" % time.strftime("%Y%m%d_%H%M%S"))
    try:
        with open(csv_path, "w", newline="", encoding="utf-8-sig") as f:
            w = csv.writer(f)
            w.writerow(_REPORT_HEADER)
            for r in records:
                w.writerow([r["idx"], r["key"], r["study_uid"], r["status"],
                            r["n_files"], r["message"], r["time"]])
        return csv_path
    except Exception as e:
        _ui_queue.put(("log", "[报告] 写入下载报告失败：%s" % e))
        return None


def _download_one(cfg, key, label, uid, idx, total):
    """下载单个 Study（不重试）。返回 (idx, status, n_files, message)。"""
    global _current_label, _current_expected_images
    _current_label = label
    _current_expected_images = 0  # 拉取前未知，由 pull_one_study 返回值/异常时回填
    n_files = 0

    # 停止优先：停止后剩余任务统一记为“停止”
    if _stop_event.is_set():
        _ui_queue.put(("log", "[%d/%d] [跳过] %s（手动停止）" % (idx, total, label)))
        _record_download(idx, key, uid, label, "停止", 0)
        return idx, "stopped", 0, "已停止"

    def _success(n, note=""):
        """成功统一出口：记录 + 返回。"""
        _record_download(idx, key, uid, label, "成功", n, note)
        return idx, "success", n, note

    # 全量拉取该 Study（前置机不支持 IMAGE/SERIES 级查询，无法精准补拉，接受个别失败）
    assoc = _make_assoc(cfg, StudyRootQueryRetrieveInformationModelMove, "C-Move", "cmove_timeout", 300)
    if assoc and assoc.is_established:
        try:
            has_error, code, n_files, subdir, n_failed, n_warned, n_total = pull_one_study(assoc, uid, cfg.local_aet, OUTPUT_ROOT, key)
            _current_expected_images = n_total
            _ui_queue.put(("image_progress", (label, n_files, n_total)))
        finally:
            _unregister_assoc(assoc)
            try:
                assoc.release()
            except Exception:
                pass
        if not has_error and n_files > 0:
            # 状态码 0x0000 且无子操作失败/警告：完全成功
            if n_failed == 0 and n_warned == 0 and code == 0x0000:
                _ui_queue.put(("log", "[%d/%d] [完成] %s -> %d 个文件 %s" % (idx, total, label, n_files, subdir)))
                return _success(n_files)
            # 子操作有失败/警告（0xB000）：按本地已落盘文件数视为成功，接受个别失败
            _ui_queue.put(("log", "[%d/%d] [完成] %s -> %d 个文件（子操作 %d 失败/%d 警告，接受） %s" % (idx, total, label, n_files, n_failed, n_warned, subdir)))
            return _success(n_files, "子操作 %d 失败/%d 警告" % (n_failed, n_warned))
        elif has_error:
            fail_msg = "状态码 0x%04X（本地文件 %d 个）" % (code, n_files)
        else:
            # 状态码成功但 0 文件：前置机没有把影像推到本机（注册信息不符/防火墙/异步未推）
            ips = "、".join(_local_ips()) or "未知"
            fail_msg = ("C-Move 状态成功但等待 30 秒后仍 0 个文件落盘。"
                        "请核对：1) 医院前置机上「%s」注册的 IP 是否为本机当前 IP（%s）；"
                        "2) 注册端口是否为 %d；3) Windows 防火墙是否放行 %d 入站。"
                        % (cfg.local_aet, ips, cfg.local_port, cfg.local_port))
    else:
        fail_msg = "连接 PACS 失败"
    _ui_queue.put(("log", "[%d/%d] [失败] %s：%s" % (idx, total, label, fail_msg)))
    _record_download(idx, key, uid, label, "失败", n_files, fail_msg)
    return idx, "failed", n_files, fail_msg


def batch_download(cfg):
    """批量下载主流程（在线程中运行）。任何退出路径（成功/报错/异常）都会复位 GUI 状态并写报告。"""
    try:
        _batch_download_inner(cfg)
    except Exception as e:
        # 提取出错位置、错误类型与原因：弹窗给简洁信息，完整堆栈写日志便于排查
        exc_type = type(e).__name__
        reason = str(e) or "(无具体错误信息)"
        tb = traceback.extract_tb(sys.exc_info()[2])
        loc = "未知位置"
        if tb:
            loc = "%s 第 %d 行" % (os.path.basename(tb[-1].filename), tb[-1].lineno)
        _ui_queue.put(("log", "[错误] 出错位置：%s\n%s" % (loc, traceback.format_exc())))
        _ui_queue.put(("error", "出错位置：%s\n错误类型：%s\n错误原因：%s" % (loc, exc_type, reason)))
    finally:
        _abort_active_assocs()
        records = _snapshot_records()
        csv_path = _write_report_csv(records, cfg.out_dir) if records else None
        _ui_queue.put(("report", (records, csv_path)))
        _ui_queue.put(("reset", None))


def _batch_download_inner(cfg):
    global OUTPUT_ROOT, _rate_limit_kbps
    global _rate_last_time, _rate_tokens, _download_start_time, _throttle_cap_logged
    # 关键：重置停止标志，避免上一次“停止”被传染到本次下载
    _stop_event.clear()
    _clear_records()

    OUTPUT_ROOT = cfg.out_dir
    _download_start_time = time.time()  # 记录本次下载开始时间
    os.makedirs(OUTPUT_ROOT, exist_ok=True)
    _rate_limit_kbps = int(getattr(cfg, "rate_limit_kbps", 0) or 0)
    _rate_last_time = 0.0
    _rate_tokens = 0.0
    _throttle_cap_logged = False  # 每次下载重置，保证限速封顶提示能在每次下载时生效
    if _rate_limit_kbps > 0:
        _ui_queue.put(("log", "已启用限速：%d KB/s" % _rate_limit_kbps))

    # 1) 读 Excel（得到 keys：StudyInstanceUID 或 patientId）
    _ui_queue.put(("log", "正在读取 Excel：%s" % cfg.excel_path))
    try:
        keys = read_study_uids(cfg.excel_path, cfg.column, cfg.sheet_name or None, fallback_to_first_col=False)
    except Exception as e:
        _ui_queue.put(("error", "读取 Excel 失败：%s" % e))
        return
    n_keys = len(keys)
    key_label = "StudyInstanceUID" if cfg.key_type == "study_uid" else "病人ID(patientId)"
    _ui_queue.put(("log", "共读取到 %d 个 %s" % (n_keys, key_label)))
    if n_keys == 0:
        _ui_queue.put(("error", "未读取到任何数据，请检查列名/Sheet 配置"))
        return

    # 2) 启动 Store SCP
    if not start_store_scp(cfg.local_aet, cfg.local_port):
        return

    # 3) 展开任务：把 keys 转成 (显示标签, StudyInstanceUID) 列表
    tasks = []
    if cfg.key_type == "patient_id":
        cfind_timeout = int(getattr(cfg, "cfind_timeout", 60) or 60)
        _ui_queue.put(("log", "模式：按病人ID查询，先 C-Find 找出每个病人的所有检查（超时 %d 秒）..." % cfind_timeout))
        assoc_find = _make_assoc(cfg, StudyRootQueryRetrieveInformationModelFind, "C-Find", "cfind_timeout", 60)
        if not assoc_find or not assoc_find.is_established:
            _ui_queue.put(("error", "连接 PACS 失败（C-Find），请检查 IP/端口/AE 及网络是否可达"))
            return
        try:
            for pi, pid in enumerate(keys, 1):
                if _stop_event.is_set():
                    break
                uids = find_studies_by_patient(assoc_find, pid)
                if not uids:
                    _ui_queue.put(("log", "  [%d/%d] patientId=%s 未查到任何检查" % (pi, n_keys, pid)))
                    continue
                _ui_queue.put(("log", "  [%d/%d] patientId=%s -> 找到 %d 个检查" % (pi, n_keys, pid, len(uids))))
                for u in uids:
                    tasks.append((str(pid), "%s/%s" % (pid, u), u))
        finally:
            _unregister_assoc(assoc_find)
            try:
                assoc_find.release()
            except Exception:
                pass
    else:
        for u in keys:
            tasks.append((str(u), str(u), u))

    total = len(tasks)
    _ui_queue.put(("log", "共需下载 %d 个检查（Study）" % total))
    if total == 0:
        _ui_queue.put(("error", "没有可下载的检查"))
        return

    # 4) 串行下载（稳态，支持间隙暂停、停止中断）
    cmove_timeout = int(getattr(cfg, "cmove_timeout", 300) or 300)
    _ui_queue.put(("log", "开始下载：串行（稳态），C-Move 超时=%d 秒" % cmove_timeout))

    ok = 0
    fail = 0
    pause_every = int(getattr(cfg, "pause_every", 0) or 0)
    pause_seconds = int(getattr(cfg, "pause_seconds", 0) or 0)

    for i, (key, label, uid) in enumerate(tasks, 1):
        if _stop_event.is_set():
            _ui_queue.put(("log", "已手动停止，剩余 %d 个未处理" % (total - i + 1)))
            break
        _ui_queue.put(("progress", (i, total)))
        _ui_queue.put(("log", "[%d/%d] 拉取 %s" % (i, total, label)))
        _idx, status, _, _ = _download_one(cfg, key, label, uid, i, total)
        if status == "success":
            ok += 1
        elif status == "stopped":
            break
        else:
            fail += 1

        # 间隙：每下载 N 个检查后暂停 M 秒（最后一个不暂停）
        if pause_every > 0 and pause_seconds > 0 and i % pause_every == 0 and i < total:
            if _stop_event.is_set():
                break
            _ui_queue.put(("status", "下载暂停中（%d 秒）..." % pause_seconds))
            _ui_queue.put(("log", "  [间隙] 已下载 %d 个，暂停 %d 秒..." % (i, pause_seconds)))
            t_end = time.time() + pause_seconds
            while time.time() < t_end and not _stop_event.is_set():
                time.sleep(0.5)
            _ui_queue.put(("status", "继续下载..."))
            _ui_queue.put(("log", "  [间隙] 暂停结束，继续下载"))

    if _stop_event.is_set():
        _ui_queue.put(("done", "已停止：成功 %d / 失败 %d / 共 %d（剩余任务未处理）" % (ok, fail, total)))
        _ui_queue.put(("log", "已停止：成功 %d / 失败 %d / 共 %d" % (ok, fail, total)))
    else:
        _ui_queue.put(("done", "下载完成：成功 %d / 失败 %d / 共 %d" % (ok, fail, total)))


# ---------------------------------------------------------------------------
# 5) GUI 界面
# ---------------------------------------------------------------------------
class App:
    def __init__(self, root):
        self.root = root
        root.title("DICOM 影像批量下载工具")
        root.geometry("760x720")
        root.minsize(720, 650)

        self.cfg = DownloadConfig()
        self.thread = None
        self.records = []       # 最近一次下载的结果记录
        self.report_csv = None  # 最近一次下载报告 CSV 路径

        self._build_widgets()
        self._load_config()
        self._poll_queue()

    # ----- 布局 -----
    def _build_widgets(self):
        pad = dict(padx=6, pady=3)

        # PACS 配置
        f1 = ttk.LabelFrame(self.root, text="PACS 前置机配置（由医院实施工程师提供）")
        f1.pack(fill="x", padx=10, pady=(10, 4))
        self._entry_pair(f1, "PACS IP", "pacs_host", default="")
        self._entry_pair(f1, "PACS 端口", "pacs_port", default="104")
        self._entry_pair(f1, "PACS AE Title", "pacs_aet", default="")

        # 本机配置
        f2 = ttk.LabelFrame(self.root, text="本机接收节点配置（需注册到 PACS 前置机）")
        f2.pack(fill="x", padx=10, pady=4)
        self._entry_pair(f2, "本机 AE Title", "local_aet", default="MYAET")
        self._entry_pair(f2, "本机接收端口", "local_port", default="11112")

        # 下载配置
        f3 = ttk.LabelFrame(self.root, text="下载配置")
        f3.pack(fill="x", padx=10, pady=4)

        row0 = ttk.Frame(f3); row0.pack(fill="x", **pad)
        ttk.Label(row0, text="Excel 文件:").pack(side="left")
        self.var_excel = tk.StringVar()
        ttk.Entry(row0, textvariable=self.var_excel, width=50).pack(side="left", padx=4)
        ttk.Button(row0, text="浏览...", command=self._browse_excel).pack(side="left")

        row_kt = ttk.Frame(f3); row_kt.pack(fill="x", **pad)
        ttk.Label(row_kt, text="查询键类型:").pack(side="left")
        self.var_key_type = tk.StringVar(value="病人ID(patientId)")
        ttk.Combobox(
            row_kt, textvariable=self.var_key_type, state="readonly", width=28,
            values=["StudyInstanceUID(影像号UID)", "病人ID(patientId)"],
        ).pack(side="left", padx=4)
        ttk.Label(row_kt, text="(选病人ID时，会先查该病人的全部检查再拉取)").pack(side="left")

        row1 = ttk.Frame(f3); row1.pack(fill="x", **pad)
        ttk.Label(row1, text="关键列名:").pack(side="left")
        self.var_column = tk.StringVar(value="影像号")
        ttk.Entry(row1, textvariable=self.var_column, width=16).pack(side="left", padx=4)
        ttk.Label(row1, text="(列名/字母A/序号1)").pack(side="left")
        ttk.Label(row1, text="  Sheet(可空):").pack(side="left", padx=(16, 0))
        self.var_sheet = tk.StringVar()
        ttk.Entry(row1, textvariable=self.var_sheet, width=12).pack(side="left", padx=4)

        row2 = ttk.Frame(f3); row2.pack(fill="x", **pad)
        ttk.Label(row2, text="输出目录:").pack(side="left")
        self.var_out = tk.StringVar()
        ttk.Entry(row2, textvariable=self.var_out, width=50).pack(side="left", padx=4)
        ttk.Button(row2, text="浏览...", command=self._browse_out).pack(side="left")

        row3 = ttk.Frame(f3); row3.pack(fill="x", **pad)
        ttk.Label(row3, text="限速(KB/s):").pack(side="left")
        self.var_rate_limit = tk.StringVar(value="0")
        ttk.Entry(row3, textvariable=self.var_rate_limit, width=9).pack(side="left", padx=4)
        ttk.Label(row3, text="(0=不限速, 如1024=1MB/s; 过低会触发封顶保护)").pack(side="left")

        ttk.Label(row3, text="  每下载").pack(side="left", padx=(16, 0))
        self.var_pause_every = tk.StringVar(value="0")
        ttk.Entry(row3, textvariable=self.var_pause_every, width=6).pack(side="left", padx=4)
        ttk.Label(row3, text="个检查后暂停").pack(side="left")
        self.var_pause_seconds = tk.StringVar(value="30")
        ttk.Entry(row3, textvariable=self.var_pause_seconds, width=6).pack(side="left", padx=4)
        ttk.Label(row3, text="秒(0=不暂停)").pack(side="left")

        row4 = ttk.Frame(f3); row4.pack(fill="x", **pad)
        ttk.Label(row4, text="C-Move超时(秒):").pack(side="left")
        self.var_cmove_timeout = tk.StringVar(value="300")
        ttk.Entry(row4, textvariable=self.var_cmove_timeout, width=8).pack(side="left", padx=4)
        ttk.Label(row4, text="C-Find超时(秒):").pack(side="left", padx=(12, 0))
        self.var_cfind_timeout = tk.StringVar(value="60")
        ttk.Entry(row4, textvariable=self.var_cfind_timeout, width=6).pack(side="left", padx=4)
        self.var_diag_log = tk.BooleanVar(value=False)
        ttk.Checkbutton(row4, text="诊断日志", variable=self.var_diag_log,
                        command=self._sync_diag).pack(side="left", padx=(12, 0))

        # 按钮区
        fbtn = ttk.Frame(self.root)
        fbtn.pack(fill="x", padx=10, pady=6)
        self.btn_test = ttk.Button(fbtn, text="检测连通", command=self._test_connectivity)
        self.btn_test.pack(side="left", padx=4)
        self.btn_start = ttk.Button(fbtn, text="开始下载", command=self._start)
        self.btn_start.pack(side="left", padx=4)
        self.btn_stop = ttk.Button(fbtn, text="停止", command=self._stop, state="disabled")
        self.btn_stop.pack(side="left", padx=4)
        ttk.Button(fbtn, text="保存配置", command=self._save_config).pack(side="left", padx=4)
        ttk.Button(fbtn, text="加载配置", command=self._load_config).pack(side="left", padx=4)
        self.btn_history = ttk.Button(fbtn, text="下载记录", command=self._show_history, state="disabled")
        self.btn_history.pack(side="left", padx=4)

        # 进度条
        self.progress = ttk.Progressbar(self.root, mode="determinate")
        self.progress.pack(fill="x", padx=10, pady=4)
        self.var_status = tk.StringVar(value="就绪")
        ttk.Label(self.root, textvariable=self.var_status).pack(anchor="w", padx=12)
        # 第二行：当前检查的影像张数 + 总张数 + 已耗时
        info_row = ttk.Frame(self.root)
        info_row.pack(fill="x", padx=12, pady=(0, 4))
        self.var_image_count = tk.StringVar(value="影像数: 0 / ?")
        ttk.Label(info_row, textvariable=self.var_image_count).pack(side="left")
        ttk.Label(info_row, text="    ").pack(side="left")
        self.var_elapsed = tk.StringVar(value="已耗时: 00:00:00")
        ttk.Label(info_row, textvariable=self.var_elapsed).pack(side="left")

        # 日志
        self.log = scrolledtext.ScrolledText(self.root, height=16, state="disabled", wrap="word")
        self.log.pack(fill="both", expand=True, padx=10, pady=(4, 10))

    def _entry_pair(self, parent, label, attr, default=""):
        r = ttk.Frame(parent); r.pack(fill="x", padx=6, pady=3)
        ttk.Label(r, text=label + ":", width=16).pack(side="left")
        var = tk.StringVar(value=default)
        ttk.Entry(r, textvariable=var, width=40).pack(side="left", padx=4)
        setattr(self, "var_" + attr, var)

    def _iter_editable_widgets(self):
        """遍历主窗口下所有 Entry/Combobox，用于下载期间统一禁用/恢复。
        （原实现用 nametowidget(StringVar._name) 是无效的：StringVar 的 _name 是
        Tcl 变量名而非 widget 路径，会静默失败，导致下载时输入框实际未被禁用。）"""

        def walk(w):
            for ch in w.winfo_children():
                if isinstance(ch, (ttk.Entry, ttk.Combobox)):
                    yield ch
                yield from walk(ch)

        yield from walk(self.root)

    # ----- 事件 -----
    def _browse_excel(self):
        p = filedialog.askopenfilename(
            title="选择影像号 Excel 文件",
            filetypes=[("Excel 文件", "*.xlsx *.xlsm"), ("所有文件", "*.*")],
        )
        if p:
            self.var_excel.set(p)

    def _browse_out(self):
        p = filedialog.askdirectory(title="选择输出目录")
        if p:
            self.var_out.set(p)

    def _collect_cfg(self):
        c = DownloadConfig()
        c.pacs_host = self.var_pacs_host.get().strip()
        try:
            c.pacs_port = int(self.var_pacs_port.get().strip() or 104)
        except ValueError:
            c.pacs_port = 104
        c.pacs_aet = self.var_pacs_aet.get().strip()
        c.local_aet = self.var_local_aet.get().strip() or "MYAET"
        try:
            c.local_port = int(self.var_local_port.get().strip() or 11112)
        except ValueError:
            c.local_port = 11112
        c.excel_path = self.var_excel.get().strip()
        c.sheet_name = self.var_sheet.get().strip()
        c.column = self.var_column.get().strip() or "影像号"
        c.key_type = "patient_id" if "病人" in self.var_key_type.get() else "study_uid"
        c.out_dir = self.var_out.get().strip()
        try:
            c.rate_limit_kbps = int(self.var_rate_limit.get().strip() or 0)
        except ValueError:
            c.rate_limit_kbps = 0
        try:
            c.pause_every = int(self.var_pause_every.get().strip() or 0)
        except ValueError:
            c.pause_every = 0
        try:
            c.pause_seconds = int(self.var_pause_seconds.get().strip() or 30)
        except ValueError:
            c.pause_seconds = 30
        try:
            c.cmove_timeout = int(self.var_cmove_timeout.get().strip() or 300)
        except ValueError:
            c.cmove_timeout = 300
        try:
            c.cfind_timeout = int(self.var_cfind_timeout.get().strip() or 60)
        except ValueError:
            c.cfind_timeout = 60
        return c

    def _test_connectivity(self):
        host = self.var_pacs_host.get().strip()
        if not host:
            messagebox.showwarning("提示", "请先填写 PACS IP")
            return
        try:
            port = int(self.var_pacs_port.get().strip() or 104)
        except ValueError:
            messagebox.showwarning("提示", "PACS 端口必须是数字")
            return
        aet = self.var_pacs_aet.get().strip() or None
        local_aet = self.var_local_aet.get().strip() or None
        try:
            local_port = int(self.var_local_port.get().strip() or 11112)
        except ValueError:
            local_port = None

        self._append_log("正在双向检测 %s:%d 连通性（正向 C-Echo + 反向接收探测）..." % (host, port))
        self.btn_test.config(state="disabled")

        def worker():
            ok, msgs = test_connectivity(host, port, aet, local_port=local_port, local_aet=local_aet)
            _ui_queue.put(("conn_result", (ok, msgs)))

        threading.Thread(target=worker, daemon=True).start()

    def _sync_diag(self):
        """勾选/取消「诊断日志」立即生效（不必等下次开始下载）。"""
        global _diag_log
        _diag_log = bool(self.var_diag_log.get())
        self._append_log("诊断日志已%s" % ("开启" if _diag_log else "关闭"))

    def _start(self):
        global _diag_log
        if self.thread and self.thread.is_alive():
            messagebox.showinfo("提示", "已有下载任务在运行中")
            return
        _diag_log = bool(self.var_diag_log.get())
        try:
            int(self.var_pacs_port.get().strip() or 104)
            int(self.var_local_port.get().strip() or 11112)
        except ValueError:
            messagebox.showwarning("提示", "PACS端口和本机接收端口必须是数字")
            return
        cfg = self._collect_cfg()
        if not cfg.pacs_host or not cfg.pacs_aet:
            messagebox.showwarning("提示", "请填写 PACS IP 和 AE Title")
            return
        # AE Title 校验：1~16 字符，不能含空格/中文/特殊字符（pynetdicom 编码会失败或被 PACS 拒）
        for nm, v in (("PACS AE Title", cfg.pacs_aet), ("本机 AE Title", cfg.local_aet)):
            if not v:
                messagebox.showwarning("提示", "%s 不能为空" % nm)
                return
            if len(v) > 16 or any(ch.isspace() for ch in v) or not v.isascii() or not v.isprintable():
                messagebox.showwarning("提示", "%s 不合法：必须是 1~16 位 ASCII 可显示字符，且不能含空格（当前值：%r）" % (nm, v))
                return
        if not cfg.excel_path or not os.path.exists(cfg.excel_path):
            messagebox.showwarning("提示", "请选择有效的影像号 Excel 文件")
            return
        if not cfg.out_dir:
            messagebox.showwarning("提示", "请选择输出目录")
            return

        self.progress["maximum"] = 100
        self.progress["value"] = 0
        self._clear_log()
        # 启动后禁用关键输入框，防止中途修改导致与后台线程不一致
        for w in self._iter_editable_widgets():
            try:
                w.config(state="disabled")
            except Exception:
                pass
        self.btn_test.config(state="disabled")
        self.btn_start.config(state="disabled")
        self.btn_stop.config(state="normal")
        self.btn_history.config(state="disabled")
        self.var_status.set("下载中...")
        self.var_image_count.set("影像数: 0 / ?")
        self.var_elapsed.set("已耗时: 00:00:00")
        # 启动 1Hz 定时器刷新"已耗时"显示
        self._download_start_ts = time.time()
        self._tick_after_id = None
        self._tick_running = True
        self._tick_update()

        self.thread = threading.Thread(target=batch_download, args=(cfg,), daemon=True)
        self.thread.start()

    def _stop(self):
        _stop_event.set()
        _abort_active_assocs()
        self._append_log("正在停止（中断当前连接，稍候即生效）...")

    def _save_config(self):
        cfg = self._collect_cfg()
        try:
            with open(CONFIG_PATH, "w", encoding="utf-8") as f:
                json.dump(cfg.__dict__, f, ensure_ascii=False, indent=2)
            self._append_log("配置已保存：%s" % CONFIG_PATH)
        except Exception as e:
            messagebox.showerror("错误", "保存配置失败：%s" % e)

    def _load_config(self):
        if not os.path.exists(CONFIG_PATH):
            return
        try:
            with open(CONFIG_PATH, "r", encoding="utf-8") as f:
                d = json.load(f)
            for k, v in d.items():
                if hasattr(self.cfg, k):
                    setattr(self.cfg, k, v)
            self.var_pacs_host.set(str(self.cfg.pacs_host or ""))
            self.var_pacs_port.set(str(self.cfg.pacs_port))
            self.var_pacs_aet.set(str(self.cfg.pacs_aet or ""))
            self.var_local_aet.set(str(self.cfg.local_aet or ""))
            self.var_local_port.set(str(self.cfg.local_port))
            self.var_out.set(str(self.cfg.out_dir or ""))
            self.var_excel.set(str(self.cfg.excel_path or ""))
            self.var_column.set(str(self.cfg.column or "影像号"))
            self.var_sheet.set(str(self.cfg.sheet_name or ""))
            if getattr(self.cfg, "key_type", "patient_id") == "patient_id":
                self.var_key_type.set("病人ID(patientId)")
            else:
                self.var_key_type.set("StudyInstanceUID(影像号UID)")
            self.var_rate_limit.set(str(getattr(self.cfg, "rate_limit_kbps", 0)))
            self.var_pause_every.set(str(getattr(self.cfg, "pause_every", 0)))
            self.var_pause_seconds.set(str(getattr(self.cfg, "pause_seconds", 30)))
            self.var_cmove_timeout.set(str(getattr(self.cfg, "cmove_timeout", 300)))
            self.var_cfind_timeout.set(str(getattr(self.cfg, "cfind_timeout", 60)))
            self._append_log("已自动加载配置：%s" % CONFIG_PATH)
        except Exception as e:
            print("load config error:", e)

    # ----- 日志/进度 -----
    def _append_log(self, text):
        self.log.config(state="normal")
        self.log.insert("end", text + "\n")
        self.log.see("end")
        self.log.config(state="disabled")

    def _clear_log(self):
        self.log.config(state="normal")
        self.log.delete("1.0", "end")
        self.log.config(state="disabled")

    def _reset_ui(self, status_text=None):
        """复位按钮/状态（下载线程结束或出错时调用），保证界面不会卡死。"""
        # 停止 1Hz 定时器
        self._tick_running = False
        if getattr(self, "_tick_after_id", None) is not None:
            try:
                self.root.after_cancel(self._tick_after_id)
            except Exception:
                pass
            self._tick_after_id = None
        # 恢复所有输入框为可编辑（下载结束/异常退出时）
        for w in self._iter_editable_widgets():
            try:
                w.config(state="normal")
            except Exception:
                pass
        self.btn_test.config(state="normal")
        self.btn_start.config(state="normal")
        self.btn_stop.config(state="disabled")
        if status_text:
            self.var_status.set(status_text)

    def _tick_update(self):
        """每秒刷新"已耗时"，下载结束后自动停止。"""
        try:
            if not getattr(self, "_tick_running", False):
                return
            start = getattr(self, "_download_start_ts", None)
            if start is not None:
                elapsed = max(0, int(time.time() - start))
                self.var_elapsed.set("已耗时: " + self._format_hms(elapsed))
            self._tick_after_id = self.root.after(1000, self._tick_update)
        except Exception:
            self._tick_after_id = None

    @staticmethod
    def _format_hms(seconds):
        s = int(seconds)
        h = s // 3600
        m = (s % 3600) // 60
        sec = s % 60
        return "%02d:%02d:%02d" % (h, m, sec)

    def _show_history(self):
        """弹出下载记录列表窗口。"""
        if not self.records:
            messagebox.showinfo("提示", "暂无下载记录，请先执行一次下载")
            return
        win = tk.Toplevel(self.root)
        win.title("下载记录")
        win.geometry("1060x520")
        win.transient(self.root)

        cols = ("idx", "key", "uid", "status", "n_files", "message", "time")
        headers = tuple(_REPORT_HEADER)  # 复用报告表头，避免重复定义
        widths = (60, 150, 240, 100, 70, 260, 150)

        main = ttk.Frame(win)
        main.pack(fill="both", expand=True, padx=8, pady=8)
        tree = ttk.Treeview(main, columns=cols, show="headings")
        for c, h, w in zip(cols, headers, widths):
            tree.heading(c, text=h)
            tree.column(c, width=w, anchor="w")
        for r in self.records:
            tree.insert("", "end", values=(
                r["idx"], r["key"], r["study_uid"], r["status"],
                r["n_files"], r["message"], r["time"],
            ))

        vsb = ttk.Scrollbar(main, orient="vertical", command=tree.yview)
        hsb = ttk.Scrollbar(main, orient="horizontal", command=tree.xview)
        tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")
        main.rowconfigure(0, weight=1)
        main.columnconfigure(0, weight=1)

        bar = ttk.Frame(win)
        bar.pack(fill="x", padx=8, pady=(0, 8))
        ttk.Button(bar, text="导出 CSV", command=lambda: self._export_history(win)).pack(side="left", padx=4)
        ttk.Button(bar, text="关闭", command=win.destroy).pack(side="left", padx=4)

    def _export_history(self, parent=None):
        """把当前下载记录导出为用户指定位置的 CSV。"""
        if not self.records:
            return
        p = filedialog.asksaveasfilename(
            parent=parent, title="导出下载记录",
            defaultextension=".csv", initialfile="下载记录.csv",
            filetypes=[("CSV 文件", "*.csv"), ("所有文件", "*.*")],
        )
        if not p:
            return
        try:
            with open(p, "w", newline="", encoding="utf-8-sig") as f:
                w = csv.writer(f)
                w.writerow(_REPORT_HEADER)
                for r in self.records:
                    w.writerow([r["idx"], r["key"], r["study_uid"], r["status"],
                                r["n_files"], r["message"], r["time"]])
            messagebox.showinfo("提示", "已导出：%s" % p)
        except Exception as e:
            messagebox.showerror("错误", "导出失败：%s" % e)

    def _poll_queue(self):
        try:
            while True:
                kind, payload = _ui_queue.get_nowait()
                try:
                    if kind == "log":
                        self._append_log(payload)
                    elif kind == "error":
                        self._append_log("[错误] " + payload)
                        self._reset_ui("出错，已停止（可修改配置后重新开始）")
                        # 延后弹窗：错误信息较短（位置/类型/原因），完整堆栈已写日志
                        self.root.after(50, lambda p=payload: messagebox.showerror("错误", p[:2000]))
                    elif kind == "conn_result":
                        ok, msgs = payload
                        for m in msgs:
                            self._append_log(m)
                        # 仅在无下载任务运行时才恢复检测按钮，避免与下载期间的禁用冲突
                        if not (self.thread and self.thread.is_alive()):
                            self.btn_test.config(state="normal")
                        if ok:
                            self.root.after(50, lambda m=msgs: messagebox.showinfo("检测结果", "\n".join(m)))
                        else:
                            self.root.after(50, lambda m=msgs: messagebox.showwarning("检测结果", "\n".join(m)))
                    elif kind == "progress":
                        done, total = payload
                        if total > 0:
                            self.progress["value"] = done * 100.0 / total
                            self.var_status.set("进度 %d / %d" % (done, total))
                    elif kind == "image_progress":
                        # (label, n_files, n_total) - 当前检查的影像张数进度
                        try:
                            _lbl, n_done, n_total = payload
                            total_str = str(n_total) if n_total > 0 else "?"
                            self.var_image_count.set("当前[%s] 影像数: %d / %s" % (_lbl, n_done, total_str))
                        except Exception:
                            pass
                    elif kind == "status":
                        self.var_status.set(payload)
                    elif kind == "done":
                        self._append_log(payload)
                        self.var_status.set(payload)
                        self._reset_ui()
                        self.root.after(50, lambda p=payload: messagebox.showinfo("完成", p))
                    elif kind == "report":
                        records, csv_path = payload
                        self.records = records or []
                        self.report_csv = csv_path
                        self.btn_history.config(state="normal" if self.records else "disabled")
                        if csv_path:
                            self._append_log("[报告] 下载报告已保存：%s" % csv_path)
                    elif kind == "reset":
                        # 下载线程任何退出路径都会发送，兜底复位界面
                        self._reset_ui()
                except Exception:
                    # 单条消息处理失败不影响后续轮询
                    traceback.print_exc()
        except queue.Empty:
            pass
        except Exception:
            traceback.print_exc()
        self.root.after(200, self._poll_queue)

    def _on_close(self):
        """主窗口关闭事件：优雅停止 Store SCP + 关联释放，避免 Windows 端口 TIME_WAIT。"""
        # 标记停止，避免正在跑的下载继续
        try:
            _stop_event.set()
        except Exception:
            pass
        # 主动 abort 所有活跃 DICOM association（复用统一实现）
        _abort_active_assocs()
        # 优雅关闭 Store SCP
        global _store_server, _store_started
        if _store_started and _store_server is not None:
            try:
                _store_server.shutdown()
            except Exception:
                pass
            _store_server = None
            _store_started = False
        self.root.destroy()


def main():
    root = tk.Tk()
    app = App(root)
    # 关键：拦截窗口关闭事件，优雅停 Store SCP，避免 Windows 端口 TIME_WAIT
    root.protocol("WM_DELETE_WINDOW", app._on_close)
    root.mainloop()


if __name__ == "__main__":
    main()